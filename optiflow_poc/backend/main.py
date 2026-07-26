"""
OptiFlow Backend (FastAPI Application)
--------------------------------------
Web service layer exposing CSV uploads, similar-item heuristics, and dispatch orders.
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Query, Body, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
import os
from typing import Optional, List, Dict, Any
from pydantic import BaseModel
from google.cloud import storage
import logging
import pandas as pd
import io
import time
import csv
import json
import datetime
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from pydantic import BaseModel

class StrategyUpdate(BaseModel):
    category_stores: Dict[str, List[str]]
    active_categories: List[str]

class StoreOverrideUpdate(BaseModel):
    overrides: Dict[str, str]

from data_models import (
    UploadStatus, AllocationItem, AllocationSummary,
    AllocationResponse, MatchType
)
from allocation_engine import (
    parse_planogram, parse_stock,
    run_allocation, WAREHOUSE_FACILITY, STORE_NAME_MAP
)
from regions import get_store_region, get_store_zone

app = FastAPI(
    title="OptiFlow Similarity Engine API",
    description="Automated inventory allocation based on attribute similarity matrices",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Compress large JSON responses (reduces 4MB -> ~400KB)
app.add_middleware(GZipMiddleware, minimum_size=1000)

class DataStore:
    def __init__(self):
        self.planogram: Optional[pd.DataFrame] = None
        self.planogram_dicts: List[Dict] = []
        self.sales_raw: Optional[pd.DataFrame] = None
        self.stock_raw: Optional[pd.DataFrame] = None
        self.warehouse_stock: Optional[pd.DataFrame] = None
        self.store_stock: Optional[pd.DataFrame] = None
        self.strategy_active_categories: List[str] = ["A++", "A+", "A", "B+", "B", "C"]
        self.strategy_store_lists: Dict[str, List[str]] = {}
        self.allocations: List[AllocationItem] = []
        self.allocations_dicts: List[Dict] = []
        self.summary: Optional[AllocationSummary] = None
        self.dashboard_all_stores_cache: Optional[Dict] = None
        self.last_run_time: Optional[float] = None
        self.last_run_at: Optional[str] = None

store = DataStore()

def set_store_planogram(df: Optional[pd.DataFrame]):
    if df is None or (isinstance(df, pd.DataFrame) and df.empty):
        store.planogram = None
        store.planogram_dicts = []
        return
    df = df.copy()
    from regions import get_store_region, get_store_zone
    if 'region' not in df.columns:
        df['region'] = df['store_name'].apply(get_store_region)
    if 'zone' not in df.columns:
        df['zone'] = df['region'].apply(get_store_zone)
    df['_uid'] = df.index
    store.planogram = df
    clean_df = df.astype(object).where(pd.notnull(df), None)
    store.planogram_dicts = clean_df.to_dict(orient="records")

LOCAL_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "local_data")
os.makedirs(LOCAL_DATA_DIR, exist_ok=True)
GCS_BUCKET_NAME = "optiflow-poc-data-kopal500607"

def download_from_gcs(filename):
    gcs_success = False
    gcs_data = None
    try:
        client = storage.Client()
        bucket = client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(filename)
        if blob.exists():
            gcs_data = blob.download_as_bytes()
            gcs_success = True
            # Cache locally
            local_path = os.path.join(LOCAL_DATA_DIR, filename)
            with open(local_path, "wb") as f:
                f.write(gcs_data)
    except Exception as e:
        logging.warning(f"Failed GCS download for {filename}: {e}. Falling back to local cache.")
    
    if gcs_success:
        return gcs_data

    # Fallback to local cache
    local_path = os.path.join(LOCAL_DATA_DIR, filename)
    if os.path.exists(local_path):
        try:
            with open(local_path, "rb") as f:
                return f.read()
        except Exception as e:
            logging.error(f"Error reading local file {local_path}: {e}")
    return None

def upload_to_gcs(filename, contents):
    # Always save locally first
    local_path = os.path.join(LOCAL_DATA_DIR, filename)
    try:
        with open(local_path, "wb") as f:
            f.write(contents)
    except Exception as e:
        logging.error(f"Failed to write local file {local_path}: {e}")

    # Try upload to GCS
    try:
        client = storage.Client()
        bucket = client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(filename)
        blob.upload_from_string(contents)
    except Exception as e:
        logging.warning(f"Error uploading {filename} to GCS (local save succeeded): {e}")

def delete_from_gcs(filename):
    # Delete locally first
    local_path = os.path.join(LOCAL_DATA_DIR, filename)
    if os.path.exists(local_path):
        try:
            os.remove(local_path)
        except Exception as e:
            logging.error(f"Failed to delete local file {local_path}: {e}")

    # Try deleting from GCS
    try:
        client = storage.Client()
        bucket = client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(filename)
        if blob.exists():
            blob.delete()
    except Exception as e:
        logging.warning(f"Error deleting {filename} from GCS (local delete succeeded): {e}")


import concurrent.futures

@app.on_event("startup")
async def startup_event():
    print("Downloading data from GCS in parallel...")
    files_to_download = [
        "last_run_metadata.json",
        "Planogram.csv",
        "planogram_edited.json",
        "Stock data.csv",
        "Sales Data.csv",
        "strategy_settings.json",
        "allocation_results.json",
        "dashboard_cache.json"
    ]
    
    downloaded_data = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=7) as executor:
        future_to_file = {executor.submit(download_from_gcs, f): f for f in files_to_download}
        for future in concurrent.futures.as_completed(future_to_file):
            filename = future_to_file[future]
            try:
                downloaded_data[filename] = future.result()
                print(f"Downloaded {filename}")
            except Exception as e:
                print(f"Failed to download {filename}: {e}")

    # Process metadata
    if downloaded_data.get("last_run_metadata.json"):
        try:
            metadata = json.loads(downloaded_data["last_run_metadata.json"].decode("utf-8"))
            store.last_run_at = metadata.get("last_run_at")
        except Exception as e:
            print(f"Failed to load metadata: {e}")

    # Process planogram
    if downloaded_data.get("planogram_edited.json"):
        try:
            # Load edited JSON
            data = json.loads(downloaded_data["planogram_edited.json"].decode("utf-8"))
            set_store_planogram(pd.DataFrame(data))
        except Exception as e:
            print(f"Failed to load planogram_edited.json: {e}")
    elif downloaded_data.get("Planogram.csv"):
        try:
            df = pd.read_csv(io.BytesIO(downloaded_data["Planogram.csv"]), encoding="utf-8", on_bad_lines="skip")
            # Double header check for planogram formats
            if "Unnamed: 0" in df.columns or "Unnamed: 1" in df.columns:
                # Try header=1 (row 2)
                df = pd.read_csv(io.BytesIO(downloaded_data["Planogram.csv"]), encoding="utf-8", header=1, on_bad_lines="skip")
                if "Unnamed: 0" in df.columns or "Unnamed: 1" in df.columns or "Store-Brand-Product Type" not in df.columns and "Store & Code" not in df.columns:
                    # Try header=2 (row 3) for the new format
                    df = pd.read_csv(io.BytesIO(downloaded_data["Planogram.csv"]), encoding="utf-8", header=2, on_bad_lines="skip")
            set_store_planogram(parse_planogram(df.copy()))
        except Exception as e:
            print(f"Failed to load Planogram: {e}")
            
    # Process stock
    if downloaded_data.get("Stock data.csv"):
        try:
            df = pd.read_csv(io.BytesIO(downloaded_data["Stock data.csv"]), encoding="utf-8", on_bad_lines="skip")
            store.stock_raw = df
            wh_stock, st_stock = parse_stock(df.copy())
            store.warehouse_stock = wh_stock
            store.store_stock = st_stock
        except Exception as e:
            print(f"Failed to load Stock: {e}")
            
    # Process sales
    if downloaded_data.get("Sales Data.csv"):
        try:
            df = pd.read_csv(io.BytesIO(downloaded_data["Sales Data.csv"]), encoding="utf-8", on_bad_lines="skip")
            store.sales_raw = df
        except Exception as e:
            print(f"Failed to load Sales: {e}")
            
    # Process strategy
    if downloaded_data.get("strategy_settings.json"):
        try:
            strategy_data = json.loads(downloaded_data["strategy_settings.json"].decode("utf-8"))
            store.strategy_active_categories = strategy_data.get("active_categories", [])
            store.strategy_store_lists = strategy_data.get("store_lists", {})
        except Exception as e:
            print(f"Failed to load strategy: {e}")

    # Process allocations
    has_saved_results = False
    if downloaded_data.get("allocation_results.json"):
        try:
            alloc_data = json.loads(downloaded_data["allocation_results.json"].decode("utf-8"))
            store.last_run_at = alloc_data.get("last_run_at")
            alloc_results_raw = alloc_data.get("results", [])
            store.allocations = [AllocationItem(**item) for item in alloc_results_raw]
            store.allocations_dicts = alloc_results_raw
            if "summary" in alloc_data and alloc_data["summary"]:
                from data_models import AllocationSummary
                store.summary = AllocationSummary(**alloc_data["summary"])
            has_saved_results = True
        except Exception as e:
            print(f"Failed to load allocation results: {e}")

    # Process cache
    if downloaded_data.get("dashboard_cache.json"):
        try:
            store.dashboard_all_stores_cache = json.loads(downloaded_data["dashboard_cache.json"].decode("utf-8"))
        except Exception as e:
            print(f"Failed to load dashboard cache: {e}")

    if store.planogram is not None and store.warehouse_stock is not None and not has_saved_results:
        try:
            print("Running initial allocation engine...")
            if not store.strategy_store_lists:
                stores_df = store.planogram[["store_name", "store_category"]].drop_duplicates()
                all_cats = ["A++", "A+", "A", "B+", "B", "C"]
                for cat in all_cats:
                    store.strategy_store_lists[cat] = []
                for _, row in stores_df.iterrows():
                    cat = row["store_category"]
                    if cat in store.strategy_store_lists:
                        if row["store_name"] not in store.strategy_store_lists[cat]:
                            store.strategy_store_lists[cat].append(row["store_name"])
            start = time.time()
            allocations, summary = run_allocation(
                planogram_df=store.planogram,
                wh_stock_df=store.warehouse_stock,
                store_stock_df=store.store_stock,
                sales_df=store.sales_raw,
                strategy_store_lists=store.strategy_store_lists,
                active_categories=store.strategy_active_categories
            )
            store.allocations = allocations
            store.allocations_dicts = [a.model_dump(mode='json') for a in allocations]
            store.summary = summary
            store.last_run_time = time.time() - start
            print("Initial allocation complete.")
        except Exception as e:
            print(f"Failed initial allocation: {e}")

@app.post("/api/upload/planogram")
async def upload_planogram(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents), encoding="utf-8", on_bad_lines="skip")
        
        # Double header check
        if "Unnamed: 0" in df.columns or "Unnamed: 1" in df.columns:
            df = pd.read_csv(io.BytesIO(contents), encoding="utf-8", header=1, on_bad_lines="skip")
            if "Unnamed: 0" in df.columns or "Unnamed: 1" in df.columns or "Store-Brand-Product Type" not in df.columns and "Store & Code" not in df.columns:
                df = pd.read_csv(io.BytesIO(contents), encoding="utf-8", header=2, on_bad_lines="skip")
        set_store_planogram(parse_planogram(df.copy()))
        
        upload_to_gcs("Planogram.csv", contents)
        # Delete any edited planogram json so the new CSV takes precedence on next restart
        delete_from_gcs("planogram_edited.json")
        
        return {
            "status": "success",
            "filename": file.filename,
            "rows": len(store.planogram),
            "stores": store.planogram["store_name"].nunique(),
            "brands": store.planogram["brand_name"].nunique()
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Planogram parse error: {str(e)}")


@app.post("/api/upload/stock")
async def upload_stock(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents), encoding="utf-8", on_bad_lines="skip")
        store.stock_raw = df
        
        wh_stock, st_stock = parse_stock(df.copy())
        store.warehouse_stock = wh_stock
        store.store_stock = st_stock
        
        upload_to_gcs("Stock data.csv", contents)
        
        return {
            "status": "success",
            "filename": file.filename,
            "total_rows": len(df),
            "warehouse_rows": len(wh_stock),
            "store_rows": len(st_stock),
            "warehouse_skus": wh_stock["item_code"].nunique() if not wh_stock.empty else 0,
            "warehouse_total_units": int(wh_stock["batch_stock"].sum()) if not wh_stock.empty else 0
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Stock data parse error: {str(e)}")


@app.post("/api/upload/sales")
async def upload_sales(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents), encoding="utf-8", on_bad_lines="skip")
        store.sales_raw = df
        
        upload_to_gcs("Sales Data.csv", contents)
        
        return {
            "status": "success",
            "filename": file.filename,
            "rows": len(df)
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Sales data parse error: {str(e)}")


@app.get("/api/upload/status")
async def upload_status():
    return UploadStatus(
        planogram=store.planogram is not None,
        sales=store.sales_raw is not None,
        stock=store.warehouse_stock is not None,
        planogram_rows=len(store.planogram) if store.planogram is not None else 0,
        sales_rows=len(store.sales_raw) if store.sales_raw is not None else 0,
        stock_rows=len(store.stock_raw) if store.stock_raw is not None else 0,
        warehouse_skus=store.warehouse_stock["item_code"].nunique() if store.warehouse_stock is not None else 0,
        last_run_at=store.last_run_at
    )


class AllocationRunRequest(BaseModel):
    sales_lookback_days: Optional[int] = None

@app.post("/api/run-allocation")
def execute_allocation(req: Optional[AllocationRunRequest] = None):
    lookback_days = req.sales_lookback_days if req else None
    if store.planogram is None:
        raise HTTPException(status_code=400, detail="Upload Planogram data first")
    if store.warehouse_stock is None:
        raise HTTPException(status_code=400, detail="Upload Stock data first")
        
    # Auto-initialize strategy lists if empty
    if not store.strategy_store_lists and store.planogram is not None:
        stores_df = store.planogram[["store_name", "store_category"]].drop_duplicates()
        all_cats = ["A", "B", "C"]
        for cat in all_cats:
            store.strategy_store_lists[cat] = []
        for _, row in stores_df.iterrows():
            cat = row["store_category"]
            if cat in store.strategy_store_lists:
                if row["store_name"] not in store.strategy_store_lists[cat]:
                    store.strategy_store_lists[cat].append(row["store_name"])
        
    start = time.time()
    
    allocations, summary = run_allocation(
        planogram_df=store.planogram,
        wh_stock_df=store.warehouse_stock,
        store_stock_df=store.store_stock,
        sales_df=store.sales_raw,
        strategy_store_lists=store.strategy_store_lists,
        active_categories=store.strategy_active_categories,
        sales_lookback_days=lookback_days
    )
    
    elapsed = time.time() - start
    store.allocations = allocations
    store.allocations_dicts = [a.model_dump(mode='json') for a in allocations]
    store.summary = summary
    store.last_run_time = elapsed
    store.last_run_at = datetime.datetime.now(datetime.timezone.utc).isoformat()
    
    # Invalidate dashboard cache
    store.dashboard_all_stores_cache = None
    try:
        delete_from_gcs("dashboard_cache.json")
    except Exception as e:
        logging.warning(f"Failed to delete dashboard cache from GCS: {e}")

    upload_to_gcs("last_run_metadata.json", json.dumps({"last_run_at": store.last_run_at}).encode("utf-8"))
    
    alloc_data = {
        "last_run_at": store.last_run_at,
        "results": [a.model_dump() for a in allocations],
        "summary": summary.model_dump() if summary else None
    }
    upload_to_gcs("allocation_results.json", json.dumps(alloc_data).encode("utf-8"))
    
    strategy_data = {
        "active_categories": store.strategy_active_categories,
        "store_lists": store.strategy_store_lists
    }
    upload_to_gcs("strategy_settings.json", json.dumps(strategy_data).encode("utf-8"))
    
    return {
        "status": "success",
        "processing_time": round(elapsed, 2),
        "summary": summary.model_dump(),
        "last_run_at": store.last_run_at
    }


@app.get("/api/planogram")
def get_planogram(
    store_name: Optional[str] = Query(None),
    brand_name: Optional[str] = Query(None),
    region: Optional[str] = Query(None),
    zone: Optional[str] = Query(None),
    store_category: Optional[str] = Query(None),
    commodity: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50000, ge=1, le=100000)
):
    if store.planogram is None:
        return {"data": [], "total": 0, "page": page, "page_size": page_size, "total_pages": 0}
        
    has_filter = any([store_name, brand_name, region, zone, store_category, commodity])
    if not has_filter and hasattr(store, 'planogram_dicts') and store.planogram_dicts:
        total = len(store.planogram_dicts)
        start = (page - 1) * page_size
        end = start + page_size
        paginated_data = store.planogram_dicts[start:end]
        return {
            "data": paginated_data,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size
        }
        
    df = store.planogram
    # Apply filters safely
    if store_name and isinstance(store_name, str):
        store_list = [s.strip().lower() for s in store_name.split(',')]
        df = df[df['store_name'].fillna('').str.lower().apply(lambda x: any(s in x for s in store_list))]
    if brand_name and isinstance(brand_name, str):
        brand_list = [b.strip().lower() for b in brand_name.split(',')]
        df = df[df['brand_name'].fillna('').str.lower().apply(lambda x: any(b in x for b in brand_list))]
    if zone and isinstance(zone, str):
        zone_list = [z.strip().lower() for z in zone.split(',')]
        df = df[df['zone'].fillna('').str.lower().isin(zone_list)]
    if region and isinstance(region, str):
        region_list = [r.strip().lower() for r in region.split(',')]
        df = df[df['region'].fillna('').str.lower().isin(region_list)]
    if store_category and isinstance(store_category, str):
        cat_list = [c.strip().lower() for c in store_category.split(',')]
        df = df[df['store_category'].fillna('').str.lower().isin(cat_list)]
    if commodity and isinstance(commodity, str):
        comm_list = [c.strip().lower() for c in commodity.split(',')]
        df = df[df['commodity'].fillna('').str.lower().isin(comm_list)]
        
    total = len(df)
    start = (page - 1) * page_size
    end = start + page_size
    paginated = df.iloc[start:end]
    paginated = paginated.astype(object).where(pd.notnull(paginated), None)
    
    return {
        "data": paginated.to_dict(orient="records"),
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }


class PlanogramUpdate(BaseModel):
    _uid: int
    updates: Dict[str, Any]

@app.post("/api/planogram/update")
async def update_planogram(updates: List[Dict[str, Any]] = Body(...)):
    if store.planogram is None:
        raise HTTPException(status_code=400, detail="No planogram data loaded")
        
    try:
        new_rows = []
        drop_indices = []
        for update in updates:
            uid = update.get("_uid")
            is_deleted = update.get("_deleted")
            
            if is_deleted:
                if uid is not None and uid in store.planogram.index:
                    drop_indices.append(uid)
            elif uid is not None and uid >= 0 and uid in store.planogram.index:
                for key, val in update.items():
                    if key != "_uid" and key in store.planogram.columns:
                        # Convert numeric fields
                        if key in ["facing", "back_stock", "soh"]:
                            val = pd.to_numeric(val, errors="coerce")
                            if pd.isna(val):
                                val = 0
                        store.planogram.at[uid, key] = val
            else:
                # Add new row
                new_row = {}
                for col in store.planogram.columns:
                    val = update.get(col, "")
                    if col in ["facing", "back_stock", "soh"]:
                        val = pd.to_numeric(val, errors="coerce")
                        if pd.isna(val): val = 0
                    new_row[col] = val
                new_rows.append(new_row)
                
        if drop_indices:
            set_store_planogram(store.planogram.drop(index=drop_indices))
            
        if new_rows:
            new_df = pd.DataFrame(new_rows)
            start_idx = int(store.planogram.index.max() + 1) if len(store.planogram) > 0 else 0
            new_df.index = range(start_idx, start_idx + len(new_df))
            set_store_planogram(pd.concat([store.planogram, new_df]))
            
        # Save to local persistent file
        out_json = store.planogram.to_json(orient="records")
        upload_to_gcs("planogram_edited.json", out_json.encode("utf-8"))
        
        return {"status": "success", "message": f"Updated {len(updates)} rows"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
@app.get("/api/allocation/results")
async def get_results(
    store_name: Optional[str] = Query(None),
    brand_name: Optional[str] = Query(None),
    match_type: Optional[str] = Query(None),
    region: Optional[str] = Query(None),
    zone: Optional[str] = Query(None),
    store_category: Optional[str] = Query(None),
    commodity: Optional[str] = Query(None),
    group_by: Optional[str] = Query('none'),
    dispatch_only: bool = Query(False),
    page: int = Query(1, ge=1),
    page_size: int = Query(50000, ge=1, le=100000)
):
    results = store.allocations
    if store_name:
        store_list = [s.strip().lower() for s in store_name.split(',')]
        results = [a for a in results if any(s in a.store_name.lower() for s in store_list)]
    if brand_name:
        brand_list = [b.strip().lower() for b in brand_name.split(',')]
        results = [a for a in results if any(b in a.brand_name.lower() for b in brand_list)]
    if match_type:
        match_list = [m.strip().lower() for m in match_type.split(',')]
        results = [a for a in results if a.match_type.value in match_list]
    if zone:
        zone_list = [z.strip().lower() for z in zone.split(',')]
        results = [a for a in results if a.zone and a.zone.lower() in zone_list]
    if region:
        region_list = [r.strip().lower() for r in region.split(',')]
        results = [a for a in results if get_store_region(a.store_name).lower() in region_list]
    if store_category:
        cat_list = [c.strip().lower() for c in store_category.split(',')]
        results = [a for a in results if a.store_category.lower() in cat_list]
    if commodity:
        comm_list = [c.strip().lower() for c in commodity.split(',')]
        results = [a for a in results if a.commodity.lower() in comm_list]
    if dispatch_only:
        results = [a for a in results if a.allocated_qty > 0]
        
    total = len(results)
    start = (page - 1) * page_size
    end = start + page_size
    paginated = results[start:end]
    
    if len(results) == len(store.allocations) and hasattr(store, 'allocations_dicts') and store.allocations_dicts and len(store.allocations_dicts) == total:
        paginated_dicts = store.allocations_dicts[start:end]
    else:
        paginated_dicts = [a.model_dump(mode='json') for a in paginated]

    response = {
        "allocations": paginated_dicts,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size
    }
    
    if group_by != 'none':
        def _summarize(items):
            unique_gaps = {a.gap_id: a.deficit for a in items}
            unique_expected = {a.gap_id: (a.facing + a.back_stock) for a in items}
            unique_soh = {a.gap_id: a.current_soh for a in items}
            
            total_allocated = sum(a.allocated_qty for a in items)
            total_value = sum(a.allocated_qty * a.mrp for a in items)
            initial_deficit = sum(unique_gaps.values())
            remaining_deficit = max(0, initial_deficit - total_allocated)
            expected_total = sum(unique_expected.values())
            total_soh = sum(unique_soh.values())
            
            exact = sum(a.allocated_qty for a in items if a.match_type.value == "exact")
            similar = sum(a.allocated_qty for a in items if a.match_type.value == "similar")
            fallback = sum(a.allocated_qty for a in items if a.match_type.value == "substitute")
            allocated_skus = set(a.allocated_item_code for a in items if a.allocated_qty > 0 and a.allocated_item_code)
            uniq_pct = int((len(allocated_skus) / total_allocated) * 100) if total_allocated > 0 else 0
            
            return {
                "lines": len(unique_gaps),
                "expected": int(expected_total),
                "initial_deficit": initial_deficit,
                "allocated_qty": total_allocated,
                "remaining_deficit": remaining_deficit,
                "total_value": total_value,
                "exact_qty": exact,
                "similar_qty": similar,
                "fallback_qty": fallback,
                "uniqueness_pct": uniq_pct,
                "total_soh": int(total_soh)
            }
            
        grouped = {}
        if group_by == 'zone':
            for a in results:
                zn = a.zone or "Unassigned Zone"
                reg = a.region or "Unassigned Region"
                grd = a.store_category or "Unassigned Grade"
                st = a.store_name
                if zn not in grouped: grouped[zn] = {"items": [], "regions": {}}
                grouped[zn]["items"].append(a)
                if reg not in grouped[zn]["regions"]: grouped[zn]["regions"][reg] = {"items": [], "grades": {}}
                grouped[zn]["regions"][reg]["items"].append(a)
                if grd not in grouped[zn]["regions"][reg]["grades"]: grouped[zn]["regions"][reg]["grades"][grd] = {"items": [], "stores": {}}
                grouped[zn]["regions"][reg]["grades"][grd]["items"].append(a)
                if st not in grouped[zn]["regions"][reg]["grades"][grd]["stores"]: grouped[zn]["regions"][reg]["grades"][grd]["stores"][st] = {"items": [], "rows": []}
                grouped[zn]["regions"][reg]["grades"][grd]["stores"][st]["items"].append(a)
            
            for a in paginated:
                zn = a.zone or "Unassigned Zone"
                reg = a.region or "Unassigned Region"
                grd = a.store_category or "Unassigned Grade"
                st = a.store_name
                grouped[zn]["regions"][reg]["grades"][grd]["stores"][st]["rows"].append(a.model_dump())
                
            for zn, z_data in grouped.items():
                z_data["summary"] = _summarize(z_data["items"])
                del z_data["items"]
                for reg, r_data in z_data["regions"].items():
                    r_data["summary"] = _summarize(r_data["items"])
                    del r_data["items"]
                    for grd, g_data in r_data["grades"].items():
                        g_data["summary"] = _summarize(g_data["items"])
                        del g_data["items"]
                        for st, s_data in g_data["stores"].items():
                            s_data["summary"] = _summarize(s_data["items"])
                            del s_data["items"]
                            
        elif group_by == 'store':
            for a in results:
                st = a.store_name
                br = a.brand_name
                if st not in grouped: grouped[st] = {"items": [], "brands": {}}
                grouped[st]["items"].append(a)
                if br not in grouped[st]["brands"]: grouped[st]["brands"][br] = {"items": [], "rows": []}
                grouped[st]["brands"][br]["items"].append(a)
                
            for a in paginated:
                st = a.store_name
                br = a.brand_name
                grouped[st]["brands"][br]["rows"].append(a.model_dump())
                
            for st, s_data in grouped.items():
                s_data["summary"] = _summarize(s_data["items"])
                del s_data["items"]
                for br, b_data in s_data["brands"].items():
                    b_data["summary"] = _summarize(b_data["items"])
                    del b_data["items"]
                    
        elif group_by == 'brand':
            for a in results:
                br = a.brand_name
                if br not in grouped: grouped[br] = {"items": [], "rows": []}
                grouped[br]["items"].append(a)
                
            for a in paginated:
                br = a.brand_name
                grouped[br]["rows"].append(a.model_dump())
                
            for br, b_data in grouped.items():
                b_data["summary"] = _summarize(b_data["items"])
                del b_data["items"]
                
        response["grouped"] = grouped
    
    return Response(content=json.dumps(response), media_type="application/json")
@app.get("/api/allocation/results/export")
async def export_results(
    store_name: Optional[str] = Query(None),
    brand_name: Optional[str] = Query(None),
    match_type: Optional[str] = Query(None),
    region: Optional[str] = Query(None),
    zone: Optional[str] = Query(None),
    store_category: Optional[str] = Query(None),
    group_by: Optional[str] = Query('none'),
    dispatch_only: bool = Query(False)
):
    results = store.allocations
    if dispatch_only:
        results = [a for a in results if a.allocated_qty > 0]
        
    if store_name:
        store_list = [s.strip().lower() for s in store_name.split(',')]
        results = [a for a in results if any(s in a.store_name.lower() for s in store_list)]
    if brand_name:
        brand_list = [b.strip().lower() for b in brand_name.split(',')]
        results = [a for a in results if any(b in a.brand_name.lower() for b in brand_list)]
    if match_type:
        match_list = [m.strip().lower() for m in match_type.split(',')]
        results = [a for a in results if a.match_type.value in match_list]
    if zone:
        zone_list = [z.strip().lower() for z in zone.split(',')]
        results = [a for a in results if a.zone and a.zone.lower() in zone_list]
    if region:
        region_list = [r.strip().lower() for r in region.split(',')]
        results = [a for a in results if get_store_region(a.store_name).lower() in region_list]
    if store_category:
        cat_list = [c.strip().lower() for c in store_category.split(',')]
        results = [a for a in results if a.store_category.lower() in cat_list]
        
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation Report"
    
    if dispatch_only:
        headers = ["Allocated SKU", "Item Description", "Store Name", "Target Brand", "Dispatch Qty"]
    else:
        headers = [
            "Store Name", "Region", "Store Grade", "Target Brand", 
            "Allocated Item Code", "Allocated Item Name", "Match Type", "OptiFlow Reasoning", 
            "MRP", "Expected Stock", "Stock In Hand", "Post-Distributed Stock", 
            "Deficit", "Fulfilled Qty", "Out Of Stock",
            "Initial Warehouse Stock", "Remaining Warehouse Stock"
        ]
    
    hdr_fill = PatternFill(start_color="FFE2E8F0", end_color="FFE2E8F0", fill_type="solid")
    zone_fill = PatternFill(start_color="FFEBE4F6", end_color="FFEBE4F6", fill_type="solid")
    reg_fill = PatternFill(start_color="FFE6F2FA", end_color="FFE6F2FA", fill_type="solid")
    grd_fill = PatternFill(start_color="FFE6F9F0", end_color="FFE6F9F0", fill_type="solid")
    st_fill = PatternFill(start_color="FFFDF7E2", end_color="FFFDF7E2", fill_type="solid")
    br_fill = PatternFill(start_color="FFF8F9FA", end_color="FFF8F9FA", fill_type="solid")
    
    def apply_style(row_idx, fill, is_bold=False, is_italic=False):
        for cell in ws[row_idx]:
            cell.fill = fill
            if is_bold or is_italic:
                cell.font = Font(bold=is_bold, italic=is_italic)
                
    def write_data_row(a):
        if dispatch_only:
            ws.append([
                a.allocated_item_code or "", a.allocated_item_name or "",
                a.store_name, a.brand_name, a.allocated_qty
            ])
        else:
            expected = int(a.facing + a.back_stock)
            ws.append([
                a.store_name, a.region, a.store_category, a.brand_name,
                a.allocated_item_code or "", a.allocated_item_name or "",
                a.match_type.value.upper(), a.match_reason, a.mrp, 
                expected, int(a.current_soh), int(a.current_soh + a.allocated_qty),
                int(a.initial_gap), a.allocated_qty, int(a.remaining_gap),
                a.initial_wh_stock, a.remaining_wh_stock
            ])
        
    def write_header(text, rows, fill):
        row = [""] * (5 if dispatch_only else 21)
        row[0] = text
        if rows:
            total_allocated = sum(r.allocated_qty for r in rows)
            if dispatch_only:
                row[1] = f"Total Items to Pick: {total_allocated}"
            else:
                unique_gaps = {}
                unique_expected = {}
                total_value = 0
                exact = 0
                similar = 0
                fallback = 0
                allocated_skus = set()
                for r in rows:
                    unique_gaps[r.gap_id] = r.deficit
                    unique_expected[r.gap_id] = r.facing + r.back_stock
                    total_value += r.allocated_qty * r.mrp
                    if r.match_type.value == "exact": exact += r.allocated_qty
                    elif r.match_type.value == "similar": similar += r.allocated_qty
                    elif r.match_type.value == "substitute": fallback += r.allocated_qty
                    
                    if r.allocated_qty > 0 and r.allocated_item_code:
                        allocated_skus.add(r.allocated_item_code)
                        
                initial_deficit = sum(unique_gaps.values())
                expected_total = sum(unique_expected.values())
                remaining_deficit = max(0, initial_deficit - total_allocated)
                uniq_pct = int((len(allocated_skus) / total_allocated) * 100) if total_allocated > 0 else 0
                
                row[1] = f"Lines: {len(unique_gaps)}"
                row[2] = f"Expected: {int(expected_total)}"
                row[3] = f"Deficit: {int(initial_deficit)}"
                row[4] = f"Fulfilled: {total_allocated}"
                row[5] = f"Out of Stock: {int(remaining_deficit)}"
                row[6] = f"Total Value: ₹{int(total_value):,}"
                row[7] = f"Exact Matches: {exact}"
                row[8] = f"Similar Matches: {similar}"
                row[9] = f"Fallback Matches: {fallback}"
                row[10] = f"Uniqueness: {uniq_pct}%"
        ws.append(row)
        apply_style(ws.max_row, fill, is_bold=True)
        
    if group_by == 'region' or group_by == 'zone':
        groups = {}
        for a in results:
            zn = a.zone or "Unassigned Zone"
            reg = a.region or "Unassigned Region"
            grd = a.store_category or "Unassigned Grade"
            st = a.store_name
            br = a.brand_name
            if zn not in groups: groups[zn] = {}
            if reg not in groups[zn]: groups[zn][reg] = {}
            if grd not in groups[zn][reg]: groups[zn][reg][grd] = {}
            if st not in groups[zn][reg][grd]: groups[zn][reg][grd][st] = {}
            if br not in groups[zn][reg][grd][st]: groups[zn][reg][grd][st][br] = []
            groups[zn][reg][grd][st][br].append(a)
            
        for zn, regions in groups.items():
            zn_rows = [r for grades in regions.values() for stores in grades.values() for brands in stores.values() for br_rows in brands.values() for r in br_rows]
            write_header(f"ZONE: {zn}", zn_rows, zone_fill)
            for reg, grades in regions.items():
                reg_rows = [r for grd in grades.values() for st in grd.values() for br in st.values() for r in br]
                write_header(f"  REGION: {reg}", reg_rows, reg_fill)
                for grd, stores in grades.items():
                    grd_rows = [r for st in stores.values() for br in st.values() for r in br]
                    write_header(f"    GRADE: {grd}", grd_rows, grd_fill)
                    for st, brands in stores.items():
                        st_rows = [r for br in brands.values() for r in br]
                        write_header(f"      STORE: {st}", st_rows, st_fill)
                        for br, rows in brands.items():
                            write_header(f"        BRAND: {br}", rows, br_fill)
                            ws.append(headers)
                            apply_style(ws.max_row, hdr_fill, is_bold=True)
                            for r in rows:
                                write_data_row(r)
                            ws.append([""] * (5 if dispatch_only else 16)) # Empty row spacing
                    
    elif group_by == 'store':
        groups = {}
        for a in results:
            st = a.store_name
            br = a.brand_name
            if st not in groups: groups[st] = {}
            if br not in groups[st]: groups[st][br] = []
            groups[st][br].append(a)
            
        for st, brands in groups.items():
            st_rows = [r for br in brands.values() for r in br]
            write_header(f"STORE: {st}", st_rows, st_fill)
            for br, rows in brands.items():
                write_header(f"  BRAND: {br}", rows, br_fill)
                ws.append(headers)
                apply_style(ws.max_row, hdr_fill, is_bold=True)
                for r in rows:
                    write_data_row(r)
                ws.append([""] * (5 if dispatch_only else 16))
                
    elif group_by == 'brand':
        groups = {}
        for a in results:
            br = a.brand_name
            if br not in groups: groups[br] = []
            groups[br].append(a)
            
        for br, rows in groups.items():
            write_header(f"BRAND: {br}", rows, br_fill)
            ws.append(headers)
            apply_style(ws.max_row, hdr_fill, is_bold=True)
            for r in rows:
                write_data_row(r)
            ws.append([""] * (5 if dispatch_only else 16))
            
    else:
        ws.append(headers)
        apply_style(ws.max_row, hdr_fill, is_bold=True)
        for a in results:
            write_data_row(a)
            
    # Auto-adjust column widths slightly
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
        
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=OptiFlow_Allocation_Report.xlsx"}
    )


@app.get("/api/dispatch/export")
async def export_dispatch(
    store_name: Optional[str] = Query(None),
    brand_name: Optional[str] = Query(None),
    match_type: Optional[str] = Query(None),
    region: Optional[str] = Query(None),
    zone: Optional[str] = Query(None),
    store_category: Optional[str] = Query(None),
    group_by: Optional[str] = Query('none')
):
    results = store.allocations
    if store_name:
        results = [a for a in results if store_name.lower() in a.store_name.lower()]
    if brand_name:
        results = [a for a in results if brand_name.lower() in a.brand_name.lower()]
    if match_type:
        results = [a for a in results if a.match_type.value == match_type.lower()]
    if zone:
        results = [a for a in results if a.zone and a.zone.lower() == zone.lower()]
    if region:
        results = [a for a in results if get_store_region(a.store_name).lower() == region.lower()]
    if store_category:
        results = [a for a in results if a.store_category.lower() == store_category.lower()]
        
    # ONLY include valid allocations for dispatch
    results = [a for a in results if a.allocated_qty > 0]
        
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = ["Allocated Barcode", "Allocated Item Code", "Allocated Item Name", "Allocated Qty", "MRP"]
    
    def write_data_row(a):
        writer.writerow([
            a.allocated_barcode or "", a.allocated_item_code or "", a.allocated_item_name or "",
            a.allocated_qty, a.mrp
        ])

    def write_header(text, rows=None):
        row = [""] * 5
        row[0] = text
        if rows:
            unique_gaps = {}
            total_allocated = 0
            total_value = 0
            for r in rows:
                unique_gaps[r.gap_id] = r.deficit
                total_allocated += r.allocated_qty
                total_value += r.allocated_qty * r.mrp
            initial_deficit = sum(unique_gaps.values())
            remaining_deficit = max(0, initial_deficit - total_allocated)
            row[1] = f"Summary ➔ Initial Deficit: {int(initial_deficit)} | Allocated Qty: {total_allocated} | Remaining Deficit: {int(remaining_deficit)} | Total Value: ₹{int(total_value):,}"
        writer.writerow(row)
        
    if group_by == 'region' or group_by == 'zone':
        groups = {}
        for a in results:
            zn = a.zone or "Unassigned Zone"
            reg = a.region or "Unassigned Region"
            grd = a.store_category or "Unassigned Grade"
            st = a.store_name
            if zn not in groups: groups[zn] = {}
            if reg not in groups[zn]: groups[zn][reg] = {}
            if grd not in groups[zn][reg]: groups[zn][reg][grd] = {}
            if st not in groups[zn][reg][grd]: groups[zn][reg][grd][st] = []
            groups[zn][reg][grd][st].append(a)
            
        for zn, regions in groups.items():
            zn_rows = [r for grades in regions.values() for stores in grades.values() for st_rows in stores.values() for r in st_rows]
            write_header(f"ZONE: {zn}", zn_rows)
            for reg, grades in regions.items():
                reg_rows = [r for grd in grades.values() for st in grd.values() for r in st]
                write_header(f"  REGION: {reg}", reg_rows)
                for grd, stores in grades.items():
                    grd_rows = [r for st in stores.values() for r in st]
                    write_header(f"    GRADE: {grd}", grd_rows)
                    for st, rows in stores.items():
                        write_header(f"      STORE: {st}", rows)
                        writer.writerow(headers)
                        for r in rows:
                            write_data_row(r)
                        writer.writerow([""] * 4)
                    
    elif group_by == 'store':
        groups = {}
        for a in results:
            st = a.store_name
            br = a.brand_name
            if st not in groups: groups[st] = {}
            if br not in groups[st]: groups[st][br] = []
            groups[st][br].append(a)
            
        for st, brands in groups.items():
            st_rows = [r for br in brands.values() for r in br]
            write_header(f"STORE: {st}", st_rows)
            for br, rows in brands.items():
                write_header(f"  BRAND: {br}", rows)
                writer.writerow(headers)
                for r in rows:
                    write_data_row(r)
                writer.writerow([""] * 4)
                
    elif group_by == 'brand':
        groups = {}
        for a in results:
            br = a.brand_name
            if br not in groups: groups[br] = []
            groups[br].append(a)
            
        for br, rows in groups.items():
            write_header(f"BRAND: {br}", rows)
            writer.writerow(headers)
            for r in rows:
                write_data_row(r)
            writer.writerow([""] * 4)
            
    else:
        writer.writerow(headers)
        for a in results:
            write_data_row(a)
    return StreamingResponse(
        iter([output.getvalue()]), 
        media_type="text/csv", 
        headers={"Content-Disposition": "attachment; filename=OptiFlow_Dispatch_Order.csv"}
    )
@app.get("/api/settings/strategy")
async def get_strategy():
    if store.planogram is None:
        return {"status": "error", "message": "No planogram uploaded"}
    
    stores_df = store.planogram[["store_name", "store_category"]].drop_duplicates()
    
    df_planogram = store.planogram.copy()
    df_planogram["soh_num"] = pd.to_numeric(df_planogram["soh"], errors="coerce").fillna(0)
    soh_agg = df_planogram.groupby("store_name")["soh_num"].sum().to_dict()
    
    sales_dict = {}
    if store.sales_raw is not None and not store.sales_raw.empty:
        df = store.sales_raw.copy()
        
        # Handle variations in column names
        qty_col = "Quantity" if "Quantity" in df.columns else "quantity" if "quantity" in df.columns else None
        fac_col = "Facility Name" if "Facility Name" in df.columns else "Branch Name" if "Branch Name" in df.columns else None
        
        if qty_col and fac_col:
            df["quantity_parsed"] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
            sales_agg = df.groupby(fac_col)["quantity_parsed"].sum().to_dict()
        else:
            sales_agg = {}
            
        for st_name in stores_df["store_name"]:
            mapped_facility = STORE_NAME_MAP.get(st_name, st_name)
            total_6m = sales_agg.get(mapped_facility, 0)
            total_soh = soh_agg.get(st_name, 0)
            
            # STR = Total Sales / (Current SOH + Total Sales)
            total_inventory = total_soh + total_6m
            str_pct = (total_6m / total_inventory * 100) if total_inventory > 0 else 0
            
            sales_dict[st_name] = {
                "sales_60d": int(total_6m * (60/180)),
                "sales_30d": int(total_6m * (30/180)),
                "sales_7d": int(total_6m * (7/180)),
                "soh": int(total_soh),
                "str_pct": round(str_pct, 1)
            }
            
    # Initialize if empty
    all_cats = ["A", "B", "C"]
    if not store.strategy_store_lists:
        for cat in all_cats:
            store.strategy_store_lists[cat] = []
        for _, row in stores_df.iterrows():
            cat = row["store_category"]
            if cat in store.strategy_store_lists:
                if row["store_name"] not in store.strategy_store_lists[cat]:
                    store.strategy_store_lists[cat].append(row["store_name"])
                    
    # Handle new stores from re-uploads
    existing_in_lists = set(s for lst in store.strategy_store_lists.values() for s in lst)
    for _, row in stores_df.iterrows():
        if row["store_name"] not in existing_in_lists:
            cat = row["store_category"]
            if cat in store.strategy_store_lists:
                store.strategy_store_lists[cat].append(row["store_name"])

    # Build the structured return
    columns = {}
    for cat in all_cats:
        store_objs = []
        for sname in store.strategy_store_lists.get(cat, []):
            sd = sales_dict.get(sname, {"sales_60d": 0, "sales_30d": 0, "sales_7d": 0, "soh": 0, "str_pct": 0})
            store_objs.append({
                "store_name": sname,
                "category": cat,
                "sales_7d": sd["sales_7d"],
                "sales_30d": sd["sales_30d"],
                "sales_60d": sd["sales_60d"],
                "soh": sd["soh"],
                "str_pct": sd["str_pct"]
            })
        columns[cat] = store_objs

    return {
        "status": "success",
        "categories": all_cats,
        "active_categories": store.strategy_active_categories,
        "columns": columns
    }

@app.post("/api/settings/strategy")
async def update_strategy(req: StrategyUpdate):
    store.strategy_active_categories = req.active_categories
    store.strategy_store_lists = req.category_stores
    store.dashboard_all_stores_cache = None

    # Sync store grade changes to planogram master data if loaded
    if store.planogram is not None and req.category_stores:
        try:
            for cat, store_list in req.category_stores.items():
                for s_name in store_list:
                    mask = store.planogram['store_name'].fillna('').str.lower() == s_name.lower()
                    store.planogram.loc[mask, 'store_category'] = cat
            
            out_json = store.planogram.to_json(orient="records")
            upload_to_gcs("planogram_edited.json", out_json.encode("utf-8"))
            if hasattr(store, 'planogram_dicts'):
                store.planogram_dicts = store.planogram.to_dict(orient="records")
        except Exception as e:
            logging.warning(f"Failed to sync planogram store grades: {e}")

    strategy_data = {
        "active_categories": store.strategy_active_categories,
        "store_lists": store.strategy_store_lists
    }
    upload_to_gcs("strategy_settings.json", json.dumps(strategy_data).encode("utf-8"))
    
    return {"status": "success"}


@app.get("/api/allocation/summary")
async def get_summary():
    if store.summary is None:
        return {"status": "no_data", "summary": None}
        
    store.summary.last_run_at = store.last_run_at
    
    return {
        "status": "ready",
        "summary": store.summary.model_dump(),
        "processing_time": store.last_run_time,
        "last_run_at": store.last_run_at
    }


@app.get("/api/allocation/store-detail/{store_name}")
async def get_store_detail(store_name: str):
    if not store.allocations:
        raise HTTPException(status_code=404, detail="Run the allocation engine first")
        
    store_allocs = [
        a for a in store.allocations
        if a.store_name.lower() == store_name.lower()
    ]
    if not store_allocs:
        raise HTTPException(status_code=404, detail=f"No allocations for store: {store_name}")
        
    brand_comm_deficits = {}
    brand_comm_soh = {}
    for a in store_allocs:
        key = a.gap_id
        brand_comm_deficits[key] = max(0, a.deficit)
        brand_comm_soh[key] = max(0, a.current_soh)
        
    total_deficit = int(sum(brand_comm_deficits.values()))
    total_soh = int(sum(brand_comm_soh.values()))

    total_filled = sum(a.allocated_qty for a in store_allocs)
    total_retail_value = sum(a.allocated_qty * a.mrp for a in store_allocs)
    exact_count = sum(1 for a in store_allocs if a.match_type == MatchType.EXACT)
    total_lines = len([a for a in store_allocs if a.match_type != MatchType.UNRESOLVED])
    
    brands = {}
    for a in store_allocs:
        if a.brand_name not in brands:
            brands[a.brand_name] = {
                "brand_name": a.brand_name,
                "deficit": 0,
                "filled": 0,
                "out_of_stock": 0,
                "exact_lines": 0,
                "similar_lines": 0,
                "fallback_lines": 0,
                "retail_value": 0.0,
                "items": []
            }
        brands[a.brand_name]["items"].append(a.model_dump())
        brands[a.brand_name]["filled"] += a.allocated_qty
        brands[a.brand_name]["retail_value"] += a.allocated_qty * a.mrp
        
        if a.match_type == MatchType.EXACT:
            brands[a.brand_name]["exact_lines"] += 1
        elif a.match_type == MatchType.SIMILAR:
            brands[a.brand_name]["similar_lines"] += 1
        elif a.match_type == MatchType.SUBSTITUTE:
            brands[a.brand_name]["fallback_lines"] += 1
            
    gap_map = {a.gap_id: a for a in store_allocs}
    for b_name in brands:
        b_def = sum(v for k, v in brand_comm_deficits.items() if gap_map[k].brand_name == b_name)
        brands[b_name]["deficit"] = int(b_def)
        brands[b_name]["out_of_stock"] = int(max(0, brands[b_name]["deficit"] - brands[b_name]["filled"]))
        brands[b_name]["fulfillment_pct"] = round((brands[b_name]["filled"] / brands[b_name]["deficit"] * 100), 1) if brands[b_name]["deficit"] > 0 else 100
        
    return {
        "store_name": store_allocs[0].store_name,
        "store_category": store_allocs[0].store_category,
        "store_type": store_allocs[0].store_type,
        "total_deficit": int(total_deficit),
        "total_soh": int(total_soh),
        "total_filled": total_filled,
        "fulfillment_pct": round((total_filled / total_deficit * 100), 1) if total_deficit > 0 else 100,
        "total_unresolved": max(0, int(total_deficit) - total_filled),
        "total_retail_value": total_retail_value,
        "exact_matches": exact_count,
        "similar_matches": sum(1 for a in store_allocs if a.match_type == MatchType.SIMILAR),
        "brand_fallbacks": sum(1 for a in store_allocs if a.match_type == MatchType.SUBSTITUTE),
        "unresolved_lines": sum(1 for a in store_allocs if a.match_type == MatchType.UNRESOLVED),
        "total_deficit_lines": total_lines + sum(1 for a in store_allocs if a.match_type == MatchType.UNRESOLVED),
        "total_items": sum(a.allocated_qty for a in store_allocs),
        "total_lines": len(store_allocs),
        "match_accuracy_pct": round((exact_count / (exact_count + sum(1 for a in store_allocs if a.match_type == MatchType.SIMILAR) + sum(1 for a in store_allocs if a.match_type == MatchType.SUBSTITUTE)) * 100), 1) if (exact_count + sum(1 for a in store_allocs if a.match_type == MatchType.SIMILAR) + sum(1 for a in store_allocs if a.match_type == MatchType.SUBSTITUTE)) > 0 else 100,
        "uniqueness_pct": round((len(set(a.allocated_item_code for a in store_allocs if a.allocated_qty > 0 and a.allocated_item_code)) / total_filled * 100), 1) if total_filled > 0 else 0,
        "brands": list(brands.values())
    }


@app.get("/api/allocation/brand-detail/{brand_name}")
async def get_brand_detail(brand_name: str):
    if not store.allocations:
        raise HTTPException(status_code=404, detail="Run the allocation engine first")
        
    brand_allocs = [
        a for a in store.allocations
        if a.brand_name.lower() == brand_name.lower()
    ]
    if not brand_allocs:
        raise HTTPException(status_code=404, detail=f"No allocations for brand: {brand_name}")
        
    store_comm_deficits = {}
    store_comm_soh = {}
    for a in brand_allocs:
        key = a.gap_id
        store_comm_deficits[key] = max(0, a.deficit)
        store_comm_soh[key] = max(0, a.current_soh)
        
    total_deficit = int(sum(store_comm_deficits.values()))
    total_soh = int(sum(store_comm_soh.values()))

    total_filled = sum(a.allocated_qty for a in brand_allocs)
    total_retail_value = sum(a.allocated_qty * a.mrp for a in brand_allocs)
    
    exact_lines = sum(1 for a in brand_allocs if a.match_type == MatchType.EXACT)
    similar_lines = sum(1 for a in brand_allocs if a.match_type == MatchType.SIMILAR)
    fallback_lines = sum(1 for a in brand_allocs if a.match_type == MatchType.SUBSTITUTE)
    unresolved_lines = sum(1 for a in brand_allocs if a.match_type == MatchType.UNRESOLVED)
    
    stores_data = {}
    for a in brand_allocs:
        if a.store_name not in stores_data:
            stores_data[a.store_name] = {
                "store_name": a.store_name,
                "deficit": 0,
                "filled": 0,
                "out_of_stock": 0,
                "exact_lines": 0,
                "similar_lines": 0,
                "fallback_lines": 0,
                "retail_value": 0.0,
                "items": []
            }
        stores_data[a.store_name]["items"].append(a.model_dump())
        stores_data[a.store_name]["filled"] += a.allocated_qty
        stores_data[a.store_name]["retail_value"] += a.allocated_qty * a.mrp
        
        if a.match_type == MatchType.EXACT:
            stores_data[a.store_name]["exact_lines"] += 1
        elif a.match_type == MatchType.SIMILAR:
            stores_data[a.store_name]["similar_lines"] += 1
        elif a.match_type == MatchType.SUBSTITUTE:
            stores_data[a.store_name]["fallback_lines"] += 1
            
    gap_map = {a.gap_id: a for a in brand_allocs}
    for s_name in stores_data:
        s_def = sum(v for k, v in store_comm_deficits.items() if gap_map[k].store_name == s_name)
        stores_data[s_name]["deficit"] = int(s_def)
        stores_data[s_name]["out_of_stock"] = int(max(0, stores_data[s_name]["deficit"] - stores_data[s_name]["filled"]))
        stores_data[s_name]["fulfillment_pct"] = round((stores_data[s_name]["filled"] / stores_data[s_name]["deficit"] * 100), 1) if stores_data[s_name]["deficit"] > 0 else 100
        
    return {
        "brand_name": brand_allocs[0].brand_name,
        "total_deficit": int(total_deficit),
        "total_soh": int(total_soh),
        "total_filled": total_filled,
        "fulfillment_pct": round((total_filled / total_deficit * 100), 1) if total_deficit > 0 else 100,
        "total_unresolved": max(0, int(total_deficit) - total_filled),
        "total_retail_value": total_retail_value,
        "exact_matches": exact_lines,
        "similar_matches": similar_lines,
        "brand_fallbacks": fallback_lines,
        "unresolved_lines": unresolved_lines,
        "total_deficit_lines": exact_lines + similar_lines + fallback_lines + unresolved_lines,
        "total_items": sum(a.allocated_qty for a in brand_allocs),
        "total_lines": len(brand_allocs),
        "stores": list(stores_data.values())
    }


@app.get("/api/allocation/download/by-store")
async def download_csv_by_store():
    if not store.allocations:
        raise HTTPException(status_code=400, detail="No allocations to download")
    
    allocs = sorted(store.allocations, key=lambda a: (a.store_name, a.brand_name))
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Store", "Tier", "Brand", "Allocated SKU", "Allocated Barcode", "Model", "Color", "Qty", "MRP", "Match Type", "Reason", "Deficit", "SOH"])
    
    for a in allocs:
        writer.writerow([
            a.store_name, a.store_category, a.brand_name, a.allocated_item_code, a.allocated_barcode or "",
            a.allocated_attributes.model if a.allocated_attributes else "",
            a.allocated_attributes.color if a.allocated_attributes else "",
            a.allocated_qty, a.mrp, a.match_type.value, a.match_reason, a.deficit, a.current_soh
        ])
        
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=allocation_by_store.csv"}
    )

@app.get("/api/allocation/download/region/{region_name}")
async def download_csv_for_region(region_name: str):
    if not store.allocations:
        raise HTTPException(status_code=400, detail="No allocations to download")
        
    region_allocs = [a for a in store.allocations if get_store_region(a.store_name) == region_name]
    
    if not region_allocs:
        raise HTTPException(status_code=404, detail="Region not found")
        
    allocs = sorted(region_allocs, key=lambda a: (a.store_name, a.brand_name))
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Region", "Store", "Tier", "Brand", "Allocated SKU", "Allocated Barcode", "Model", "Color", "Qty", "MRP", "Match Type", "Reason", "Deficit", "SOH"])
    
    for a in allocs:
        writer.writerow([
            region_name, a.store_name, a.store_category, a.brand_name, a.allocated_item_code, a.allocated_barcode or "",
            a.allocated_attributes.model if a.allocated_attributes else "",
            a.allocated_attributes.color if a.allocated_attributes else "",
            a.allocated_qty, a.mrp, a.match_type.value, a.match_reason, a.deficit, a.current_soh
        ])
        
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="allocation_{region_name.lower().replace(" ", "_")}.csv"'}
    )

@app.get("/api/allocation/download/by-brand")
async def download_csv_by_brand():
    if not store.allocations:
        raise HTTPException(status_code=400, detail="No allocations to download")
    
    allocs = sorted(store.allocations, key=lambda a: (a.brand_name, a.store_name))
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Brand", "Store", "Tier", "Allocated SKU", "Allocated Barcode", "Model", "Color", "Qty", "MRP", "Match Type", "Reason", "Deficit", "SOH"])
    
    for a in allocs:
        writer.writerow([
            a.brand_name, a.store_name, a.store_category, a.allocated_item_code, a.allocated_barcode or "",
            a.allocated_attributes.model if a.allocated_attributes else "",
            a.allocated_attributes.color if a.allocated_attributes else "",
            a.allocated_qty, a.mrp, a.match_type.value, a.match_reason, a.deficit, a.current_soh
        ])
        
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=allocation_by_brand.csv"}
    )

@app.get("/api/allocation/dispatch/{store_name}")
async def get_dispatch(store_name: str):
    if not store.allocations:
        raise HTTPException(status_code=404, detail="Run the allocation engine first")
        
    store_allocs = [
        a for a in store.allocations
        if a.store_name.lower() == store_name.lower()
    ]
    if not store_allocs:
        raise HTTPException(status_code=404, detail=f"No allocations for store: {store_name}")
        
    brands = {}
    for a in store_allocs:
        if a.brand_name not in brands:
            brands[a.brand_name] = []
        brands[a.brand_name].append(a.model_dump())
        
    return {
        "store_name": store_name,
        "store_category": store_allocs[0].store_category,
        "store_type": store_allocs[0].store_type,
        "total_items": sum(a.allocated_qty for a in store_allocs),
        "total_lines": len(store_allocs),
        "brands": brands
    }


@app.get("/api/allocation/dispatch/brand/{brand_name}")
async def get_dispatch_by_brand(brand_name: str):
    if not store.allocations:
        raise HTTPException(status_code=404, detail="Run the allocation engine first")
        
    brand_allocs = [
        a for a in store.allocations
        if a.brand_name.lower() == brand_name.lower()
    ]
    if not brand_allocs:
        raise HTTPException(status_code=404, detail=f"No allocations for brand: {brand_name}")
        
    stores = {}
    for a in brand_allocs:
        if a.store_name not in stores:
            stores[a.store_name] = []
        stores[a.store_name].append(a.model_dump())
        
    return {
        "brand_name": brand_allocs[0].brand_name,
        "total_items": sum(a.allocated_qty for a in brand_allocs),
        "total_lines": len(brand_allocs),
        "stores": stores
    }


@app.get("/api/allocation/dispatch/region/{region_name}")
async def get_dispatch_by_region(region_name: str):
    if not store.allocations:
        raise HTTPException(status_code=404, detail="Run the allocation engine first")
        
    region_allocs = [
        a for a in store.allocations
        if get_store_region(a.store_name).lower() == region_name.lower()
    ]
    if not region_allocs:
        raise HTTPException(status_code=404, detail=f"No allocations for region: {region_name}")
        
    stores = {}
    for a in region_allocs:
        if a.store_name not in stores:
            stores[a.store_name] = []
        stores[a.store_name].append(a.model_dump())
        
    return {
        "region_name": region_name,
        "total_items": sum(a.allocated_qty for a in region_allocs),
        "total_lines": len(region_allocs),
        "stores": stores
    }



@app.get("/api/allocation/status")
async def allocation_status():
    return {
        "has_results": len(store.allocations) > 0,
        "total_allocations": len(store.allocations),
        "last_run_at": store.last_run_at
    }

@app.post("/api/upload/reset")
async def reset_data():
    store.planogram = None
    store.sales_raw = None
    store.stock_raw = None
    store.warehouse_stock = None
    store.store_stock = None
    store.strategy_store_lists = {}
    store.strategy_active_categories = ["A++", "A+", "A", "B+", "B", "C"]
    
    strategy_data = {
        "active_categories": store.strategy_active_categories,
        "store_lists": store.strategy_store_lists
    }
    upload_to_gcs("strategy_settings.json", json.dumps(strategy_data).encode("utf-8"))
    
    return {"status": "success"}

@app.get("/api/dashboard/all-stores")
async def dashboard_all_stores():
    """
    Dashboard endpoint that runs the full allocation engine with ALL stores active
    (ignores the Set Priority strategy settings). Also returns the constant total
    Corporate Office warehouse stock regardless of which items were allocated.
    """
    if store.dashboard_all_stores_cache is not None:
        return store.dashboard_all_stores_cache

    if store.planogram is None:
        return {"allocations": [], "warehouse_stock_total": 0}
    if store.warehouse_stock is None:
        return {"allocations": [], "warehouse_stock_total": 0}

    # Build all-stores strategy: every store in every tier, all tiers active
    all_cats = ["A++", "A+", "A", "B+", "B", "C"]
    stores_df = store.planogram[["store_name", "store_category"]].drop_duplicates()
    all_store_lists: Dict[str, List[str]] = {cat: [] for cat in all_cats}
    for _, row in stores_df.iterrows():
        cat = str(row["store_category"])
        sname = str(row["store_name"])
        if cat in all_store_lists and sname not in all_store_lists[cat]:
            all_store_lists[cat].append(sname)

    allocations, _ = run_allocation(
        planogram_df=store.planogram,
        wh_stock_df=store.warehouse_stock,
        store_stock_df=store.store_stock if store.store_stock is not None else pd.DataFrame(),
        sales_df=store.sales_raw,
        strategy_store_lists=all_store_lists,
        active_categories=all_cats
    )

    # Constant: total Corporate Office warehouse stock (not filtered by allocation)
    wh_stock_total = float(store.warehouse_stock["batch_stock"].sum())

    result = {
        "allocations": [a.model_dump() for a in allocations],
        "warehouse_stock_total": wh_stock_total
    }
    
    store.dashboard_all_stores_cache = result
    # Persist cache to GCS so it survives cold starts
    try:
        upload_to_gcs("dashboard_cache.json", json.dumps(result).encode("utf-8"))
        print("Saved dashboard_cache.json to GCS")
    except Exception as e:
        print(f"Failed to save dashboard cache to GCS: {e}")
    return result


@app.get("/api/stores")
async def list_stores():
    if store.planogram is None:
        return {"stores": []}
        
    stores = store.planogram.groupby("store_name").agg(
        store_category=("store_category", "first"),
        store_type=("store_type", "first"),
        brand_count=("brand_name", "nunique")
    ).reset_index()
    
    store_list = stores.to_dict("records")
    allocated_stores = set(a.store_name for a in store.allocations) if store.allocations else set()
    for s in store_list:
        s["has_allocations"] = s["store_name"] in allocated_stores
        
    return {"stores": store_list}


@app.get("/api/brands")
async def list_brands():
    if store.planogram is None:
        return {"brands": []}
        
    brands = store.planogram["brand_name"].unique().tolist()
    brands.sort()
    
    brand_list = [{"brand_name": b} for b in brands]
    allocated_brands = set(a.brand_name for a in store.allocations) if store.allocations else set()
    for b in brand_list:
        b["has_allocations"] = b["brand_name"] in allocated_brands
        
    return {"brands": brand_list}


@app.get("/api/regions")
async def list_regions():
    if not store.allocations:
        return {"regions": []}
    
    region_stats = {}
    
    for a in store.allocations:
        reg = get_store_region(a.store_name)
        if reg not in region_stats:
            region_stats[reg] = {
                "region_name": reg,
                "zone_name": get_store_zone(reg),
                "total_stores": set(),
                "total_deficit": 0,
                "total_filled": 0,
                "store_list": set(),
                "tier_fulfillment": {cat: {"target": 0, "filled": 0, "pct": 0.0} for cat in ["A++", "A+", "A", "B+", "B", "C"]}
            }
        region_stats[reg]["total_stores"].add(a.store_name)
        region_stats[reg]["store_list"].add(a.store_name)
        region_stats[reg]["total_filled"] += a.allocated_qty
        region_stats[reg]["tier_fulfillment"][a.store_category]["filled"] += a.allocated_qty
        
    for reg, stats in region_stats.items():
        reg_allocs = [a for a in store.allocations if get_store_region(a.store_name) == reg]
        store_comm_deficits = {}
        for a in reg_allocs:
            key = a.gap_id
            if key not in store_comm_deficits:
                store_comm_deficits[key] = max(0, a.deficit)
                stats["tier_fulfillment"][a.store_category]["target"] += int(max(0, a.deficit))
        
        stats["total_deficit"] = int(sum(store_comm_deficits.values()))
        stats["total_stores"] = len(stats["total_stores"])
        stats["fulfillment_pct"] = round((stats["total_filled"] / stats["total_deficit"] * 100), 1) if stats["total_deficit"] > 0 else 100
        stats["store_list"] = list(stats["store_list"])
        
        for cat in stats["tier_fulfillment"]:
            t = stats["tier_fulfillment"][cat]["target"]
            f = stats["tier_fulfillment"][cat]["filled"]
            stats["tier_fulfillment"][cat]["pct"] = round((f / t * 100), 1) if t > 0 else 0.0
        
    return {"regions": list(region_stats.values())}


@app.get("/api/allocation/region-detail/{region_name}")
async def get_region_detail(region_name: str):
    if not store.allocations:
        raise HTTPException(status_code=404, detail="Run the allocation engine first")
        
    region_allocs = [
        a for a in store.allocations
        if get_store_region(a.store_name).lower() == region_name.lower()
    ]
    if not region_allocs:
        raise HTTPException(status_code=404, detail=f"No allocations for region: {region_name}")
        
    store_comm_deficits = {}
    store_comm_soh = {}
    for a in region_allocs:
        key = a.gap_id
        store_comm_deficits[key] = max(0, a.deficit)
        store_comm_soh[key] = max(0, a.current_soh)
        
    total_deficit = int(sum(store_comm_deficits.values()))
    total_soh = int(sum(store_comm_soh.values()))

    total_filled = sum(a.allocated_qty for a in region_allocs)
    total_retail_value = sum(a.allocated_qty * a.mrp for a in region_allocs)
    
    exact_lines = sum(1 for a in region_allocs if a.match_type == MatchType.EXACT)
    similar_lines = sum(1 for a in region_allocs if a.match_type == MatchType.SIMILAR)
    fallback_lines = sum(1 for a in region_allocs if a.match_type == MatchType.SUBSTITUTE)
    unresolved_lines = sum(1 for a in region_allocs if a.match_type == MatchType.UNRESOLVED)
    
    stores_data = {}
    for a in region_allocs:
        if a.store_name not in stores_data:
            stores_data[a.store_name] = {
                "store_name": a.store_name,
                "store_category": a.store_category,
                "deficit": 0,
                "filled": 0,
                "out_of_stock": 0,
                "exact_lines": 0,
                "similar_lines": 0,
                "fallback_lines": 0,
                "retail_value": 0.0,
                "items": []
            }
        stores_data[a.store_name]["items"].append(a.model_dump())
        stores_data[a.store_name]["filled"] += a.allocated_qty
        stores_data[a.store_name]["retail_value"] += a.allocated_qty * a.mrp
        
        if a.match_type == MatchType.EXACT:
            stores_data[a.store_name]["exact_lines"] += 1
        elif a.match_type == MatchType.SIMILAR:
            stores_data[a.store_name]["similar_lines"] += 1
        elif a.match_type == MatchType.SUBSTITUTE:
            stores_data[a.store_name]["fallback_lines"] += 1
            
    gap_map = {a.gap_id: a for a in region_allocs}
    for s_name in stores_data:
        s_def = sum(v for k, v in store_comm_deficits.items() if gap_map[k].store_name == s_name)
        stores_data[s_name]["deficit"] = int(s_def)
        stores_data[s_name]["out_of_stock"] = int(max(0, stores_data[s_name]["deficit"] - stores_data[s_name]["filled"]))
        stores_data[s_name]["fulfillment_pct"] = round((stores_data[s_name]["filled"] / stores_data[s_name]["deficit"] * 100), 1) if stores_data[s_name]["deficit"] > 0 else 100
        
    return {
        "region_name": get_store_region(region_allocs[0].store_name),
        "total_deficit": int(total_deficit),
        "total_soh": int(total_soh),
        "total_filled": total_filled,
        "fulfillment_pct": round((total_filled / total_deficit * 100), 1) if total_deficit > 0 else 100,
        "total_unresolved": max(0, int(total_deficit) - total_filled),
        "total_retail_value": total_retail_value,
        "exact_matches": exact_lines,
        "similar_matches": similar_lines,
        "brand_fallbacks": fallback_lines,
        "unresolved_lines": unresolved_lines,
        "total_deficit_lines": exact_lines + similar_lines + fallback_lines + unresolved_lines,
        "match_accuracy_pct": round((exact_lines / (exact_lines + similar_lines + fallback_lines + unresolved_lines)) * 100, 1) if (exact_lines + similar_lines + fallback_lines + unresolved_lines) > 0 else 100,
        "total_items": sum(a.allocated_qty for a in region_allocs),
        "total_lines": len(region_allocs),
        "stores": list(stores_data.values())
    }


@app.get("/api/analytics/sales")
async def get_sales_analytics():
    """Compute brand sales velocity and predictive demand."""
    if store.sales_raw is None or store.sales_raw.empty:
        return {"status": "no_data", "analytics": []}

    df = store.sales_raw.copy()
    
    df["quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
    df["net_amount"] = pd.to_numeric(df["Net Amount"], errors="coerce").fillna(0)
    df["brand"] = df["Item Name"].apply(lambda n: str(n).split(" - ")[1].strip() if len(str(n).split(" - ")) > 1 else "Other")

    # Group by brand to calculate actual sales (past velocity)
    brand_sales = df.groupby("brand").agg(
        total_qty=("quantity", "sum"),
        total_revenue=("net_amount", "sum")
    ).reset_index()

    # Convert to monthly averages (data is 6-month block)
    brand_sales["total_qty"] = (brand_sales["total_qty"] / 6).round(0).astype(int)
    brand_sales["total_revenue"] = (brand_sales["total_revenue"] / 6).round(2)

    # Calculate basic linear projection for predictive demand (past sales + 15% growth trend)
    brand_sales["predicted_demand_30d"] = (brand_sales["total_qty"] * 1.15).round(0).astype(int)
    
    # Sort by sales volume
    brand_sales = brand_sales.sort_values("total_qty", ascending=False).head(8)

    total_rev = float(df["net_amount"].sum())

    return {
        "status": "ready",
        "brands": brand_sales.to_dict("records"),
        "total_revenue": total_rev,
        "monthly_revenue": total_rev / 6,
        "total_months": 6
    }


@app.get("/api/analytics/executive")
async def get_executive_analytics():
    """Compute store metrics, category shares, and deficit grids."""
    if store.sales_raw is None or store.sales_raw.empty:
        return {"status": "no_data"}

    sales_df = store.sales_raw.copy()
    sales_df["quantity"] = pd.to_numeric(sales_df["Quantity"], errors="coerce").fillna(0)
    sales_df["net_amount"] = pd.to_numeric(sales_df["Net Amount"], errors="coerce").fillna(0)

    # 1. Top Stores by Sales Revenue (convert to monthly run-rates)
    store_sales = sales_df.groupby("Facility Name").agg(
        total_qty=("quantity", "sum"),
        total_revenue=("net_amount", "sum")
    ).reset_index()
    store_sales["total_qty"] = (store_sales["total_qty"] / 6).round(0).astype(int)
    store_sales["total_revenue"] = (store_sales["total_revenue"] / 6).round(2)
    store_sales = store_sales.sort_values("total_revenue", ascending=False).head(6)

    # 2. Product Categories Distribution (Pie chart data) (convert to monthly averages)
    category_sales = sales_df.groupby("Item Category").agg(
        total_qty=("quantity", "sum"),
        total_revenue=("net_amount", "sum")
    ).reset_index()
    category_sales["total_qty"] = (category_sales["total_qty"] / 6).round(0).astype(int)
    category_sales["total_revenue"] = (category_sales["total_revenue"] / 6).round(2)
    category_sales = category_sales.sort_values("total_qty", ascending=False).head(5)

    # 3. Deficit Heatmap grid data
    deficit_grid = []
    if store.planogram is not None:
        pl = store.planogram.copy()
        pl["deficit"] = (pl["facing"] + pl["back_stock"]) - pl["soh"]
        pl["deficit"] = pl["deficit"].apply(lambda d: max(0, d))
        
        top_stores = pl.groupby("store_name")["deficit"].sum().sort_values(ascending=False).head(6).index.tolist()
        top_brands = pl.groupby("brand_name")["deficit"].sum().sort_values(ascending=False).head(5).index.tolist()

        for s in top_stores:
            for b in top_brands:
                sub_df = pl[(pl["store_name"] == s) & (pl["brand_name"] == b)]
                def_val = int(sub_df["deficit"].sum()) if not sub_df.empty else 0
                deficit_grid.append({
                    "store": s.replace("PO-", ""),
                    "brand": b,
                    "deficit": def_val
                })

    total_rev = float(sales_df["net_amount"].sum())

    return {
        "status": "ready",
        "stores": store_sales.to_dict("records"),
        "categories": category_sales.to_dict("records"),
        "deficit_grid": deficit_grid,
        "total_revenue": total_rev,
        "monthly_revenue": total_rev / 6,
        "total_months": 6
    }


@app.get("/api/analytics/assortment")
async def get_assortment_analytics():
    """Compute details for dead stock scatter, MRP histograms, weeks-of-cover, and Sankey flow."""
    if store.warehouse_stock is None or store.warehouse_stock.empty:
        return {"status": "no_data"}

    # 1. Price Histogram Buckets based on MRP in Stock
    mrp_col = pd.to_numeric(store.warehouse_stock["mrp"], errors="coerce").fillna(0)
    bins = [0, 1000, 2500, 5000, 10000, 160000]
    labels = ["<1k", "1k-2.5k", "2.5k-5k", "5k-10k", ">10k"]
    mrp_binned = pd.cut(mrp_col, bins=bins, labels=labels).value_counts()
    histogram = [{"bucket": k, "count": int(v)} for k, v in mrp_binned.items()]

    # 2. Dead-Stock Scatter Plot: SOH vs Sales velocity (quantity sold) per SKU
    # Group store stock by item_code
    stock_qty = store.store_stock.groupby("item_code")["batch_stock"].sum() if store.store_stock is not None else pd.Series()
    # Group sales qty by item_code
    sales_qty = pd.Series()
    if store.sales_raw is not None and not store.sales_raw.empty:
        sales_qty = store.sales_raw.groupby("Item Code")["Quantity"].sum()

    # Join
    scatter_df = pd.DataFrame({"soh": stock_qty, "sales": sales_qty}).fillna(0)
    scatter_df = scatter_df[(scatter_df["soh"] > 0) | (scatter_df["sales"] > 0)].head(100) # Limit to 100 points for chart readability
    scatter_data = [
        {
            "sku": idx,
            "soh": int(row["soh"]),
            "sales": int(row["sales"]),
            "status": "Dead Stock" if row["soh"] > 5 and row["sales"] == 0 else ("Stockout Risk" if row["soh"] == 0 and row["sales"] > 5 else "Healthy")
        }
        for idx, row in scatter_df.iterrows()
    ]

    # 3. Substitution Sankey flows (engine cascade metrics)
    sankey = {
        "exact": 0,
        "similar": 0,
        "fallback": 0,
        "unresolved": 0
    }
    if store.summary:
        sankey["exact"] = store.summary.exact_matches
        sankey["similar"] = store.summary.similar_matches
        sankey["fallback"] = store.summary.brand_fallbacks
        sankey["unresolved"] = store.summary.unresolved

    # 4. Weeks of Cover: Stock SOH / Sales Qty per store category
    # Mock cover weeks based on planogram SOH and total order ratios
    cover_data = []
    if store.planogram is not None:
        cover_df = store.planogram.groupby("store_category").agg(
            total_soh=("soh", "sum"),
            total_order=("total_order", "sum")
        ).reset_index()
        for _, row in cover_df.iterrows():
            ratio = round(float(row["total_soh"] / max(1, row["total_order"])) * 4.2, 1) # normalise to weeks-of-cover scale
            cover_data.append({
                "category": row["store_category"],
                "weeks": ratio
            })

    return {
        "status": "ready",
        "histogram": histogram,
        "scatter": scatter_data,
        "sankey": sankey,
        "cover": cover_data
    }


@app.get("/api/health")
async def health():
    return {"status": "ok", "engine": "OptiFlow Similar-Item Engine v1.0.0"}
